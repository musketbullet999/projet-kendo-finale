import openpyxl as op
import numpy as np
import xgboost as xgb
from scipy.stats import skew, kurtosis

#開啟訓練資料及測試資料
wb = op.load_workbook('TRAIN.xlsx')
sh_x = wb['DATASET']
sh_y = wb['LABEL']
data_unit = 200
data_num = int(sh_x.max_row/data_unit)
wb_test = op.load_workbook('TEST.xlsx')
test = wb_test['TEST']
test_label = wb_test['TEST_LABEL']
test_num = int(test.max_row/data_unit)

def get_rms(data_rms, feature_rms, index_rms, col_rms, nums):  #取得均方根值
    for i in range(nums):
        for j in range(data_unit):
            feature_rms[i][col_rms] += (data_rms[i][j][index_rms])**2
        feature_rms[i][col_rms] = (feature_rms[i][col_rms]/data_unit)**0.5

def get_median(data_med, feature_med, index_med, cols_med, nums):  #取得中位數
    temp_m = [0 for i in range(data_unit)]
    for i in range(nums):
        for j in range(data_unit):
            temp_m[j] = data_med[i][j][index_med]
        feature_med[i][cols_med] = np.median(temp_m)

def get_var(data_var, feature_var, index_var, cols_var, nums):  #取得變異數
    temp_v = [0 for i in range(data_unit)]
    for i in range(nums):
        for j in range(data_unit):
            temp_v[j] = data_var[i][j][index_var]
        feature_var[i][cols_var] = np.mean(temp_v)

def get_max_min(data_ext, feature_ext, func_ext, cols_ext, nums):  #取得最大最小值
    temp_e = [0 for i in range(data_unit)]
    for i in range(nums):
        for j in range(data_unit):
            temp_e[j] = data_ext[i][j]
        if(func_ext):
            feature_ext[i][cols_ext] = np.max(temp_e)
        else:
            feature_ext[i][cols_ext] = np.min(temp_e)
def get_median_1D(data_m1D, feature_m1D, cols_m1D, nums):  #從norm取得中位數
    temp_m1D = [0 for i in range(data_unit)]
    for i in range(nums):
        for j in range(data_unit):
            temp_m1D[j] = data_m1D[i][j]
        feature_m1D[i][cols_m1D] = np.median(temp_m1D)

def get_skew(data_s, feature_s, cols_s, nums):  #取得skew
    temp_s = [0 for i in range(data_unit)]
    for i in range(nums):
        for j in range(data_unit):
            temp_s[j] = data_s[i][j]
        feature_s[i][cols_s] = skew(temp_s)

def get_kurtosis(data_k, feature_k, cols_k, nums):  #取得kurtosis
    temp_k = [0 for i in range(data_unit)]
    for i in range(nums):
        for j in range(data_unit):
            temp_k[j] = data_k[i][j]
        feature_k[i][cols_k] = kurtosis(temp_k)

def sum_up(num_data, feat, col, storage, nums):  #取得總和
    for i in range(nums):
        for j in range(data_unit):
            if(num_data[i][j][col] >= 0):
                feat[i][storage] += num_data[i][j][col]
    
#建立資料儲存陣列
features = [[0 for i in range(27)]for j in range(data_num)]
labels = [0 for i in range(data_num)]
modified_data = [[[0 for i in range(6)] for j in range(data_unit)] for k in range(data_num)]
gyr_norm = [[0 for i in range(data_unit)]for j in range(data_num)]
accel_norm = [[0 for i in range(data_unit)]for j in range(data_num)]
temp = [0 for i in range(data_unit)]
tests = [[0 for i in range(27)]for j in range(test_num)]
tests_label = [0 for i in range(test_num)]
test_data = [[[0 for i in range(6)] for j in range(data_unit)] for k in range(test_num)]
test_gyr = [[0 for i in range(data_unit)]for j in range(test_num)]
test_acc = [[0 for i in range(data_unit)]for j in range(test_num)]



for i in range(data_num):  #將訓練資料從EXCEL取出
    for j in range(data_unit):
        for k in range(6):
            modified_data[i][j][k] = sh_x.cell(i*data_unit+j+1, k+1).value
    labels[i] = sh_y.cell(i+1, 1).value

for i in range(test_num):  #將測試資料從EXCEL取出
    for j in range(data_unit):
        for k in range(6):
            test_data[i][j][k] = test.cell(i*data_unit+j+1, k+1).value
    tests_label[i] = test_label.cell(i+1, 1).value 

#特徵提取
for i in range(6):  
    get_rms(modified_data, features, i, i, data_num)
    get_median(modified_data, features, i, i+6, data_num)
    get_var(modified_data, features, i, i+12, data_num)
    get_rms(test_data, tests, i, i, test_num)
    get_median(test_data, tests, i, i+6, test_num)
    get_var(test_data, tests, i, i+12, test_num)

#取得norm
for i in range(data_num):
    for j in range(data_unit):
        accel_norm[i][j] = (modified_data[i][j][0]**2 + modified_data[i][j][1]**2 + modified_data[i][j][2]**2)**0.5
        gyr_norm[i][j] = (modified_data[i][j][3]**2 + modified_data[i][j][4]**2 + modified_data[i][j][5]**2)**0.5

for i in range(test_num):
    for j in range(data_unit):
        test_acc[i][j] = (test_data[i][j][0]**2 + test_data[i][j][1]**2 + test_data[i][j][2]**2)**0.5
        test_gyr[i][j] = (test_data[i][j][3]**2 + test_data[i][j][4]**2 + test_data[i][j][5]**2)**0.5
    
get_max_min(gyr_norm, features, 0, 18, data_num)
get_max_min(gyr_norm, features, 1, 19, data_num)
get_median_1D(accel_norm, features, 20, data_num)
get_median_1D(gyr_norm, features, 21, data_num)
get_skew(accel_norm, features, 22, data_num)
get_skew(gyr_norm, features, 23, data_num)
get_kurtosis(accel_norm, features, 24, data_num)
get_kurtosis(gyr_norm, features, 25, data_num)
sum_up(modified_data, features, 4, 26, data_num)

get_max_min(test_gyr, tests, 0, 18, test_num)
get_max_min(test_gyr, tests, 1, 19, test_num)
get_median_1D(test_acc, tests, 20, test_num)
get_median_1D(test_gyr, tests, 21, test_num)
get_skew(test_acc, tests, 22, test_num)
get_skew(test_gyr, tests, 23, test_num)
get_kurtosis(test_acc, tests, 24, test_num)
get_kurtosis(test_gyr, tests, 25, test_num)
sum_up(test_data, tests, 4, 26, test_num)



#訓練及測試
x_train = np.array(features)  #將資料變換成numpy矩陣
y_train = np.array(labels).reshape(-1, 1)
x_test = np.array(tests)
y_test = np.array(tests_label).reshape(-1, 1)
dtrain = xgb.DMatrix(data=x_train, label=y_train)  #將上述矩陣轉換成DMatrix
dtest = xgb.DMatrix(data=x_test, label=y_test)

#調整參數
param = {'max_depth' : 3,            
         'eta' : 0.3,
         'objective' : 'multi:softmax',
         'num_class' : 5,
         'eval_metric':'mlogloss',
         'gamma':0.01}
num_round = 200  #設定訓練次數

#輸出訓練結果
bst = xgb.train(param, dtrain, num_round)
y_pred = bst.predict(xgb.DMatrix(x_test))
acc = 0
for d in range(len(y_pred)):
    if(y_pred[d] == y_test[d]):
        acc += 1
    else:
        print(d)
        print("Test label = ", int(tests_label[d]))
        print("Pred label = ", int(y_pred[d]))
        print("\n")
print("Accuracy = ", 100*float(acc/len(y_pred)), "%. \n")
bst.save_model('test.model')
bst.dump_model('dump.raw.txt')




    
