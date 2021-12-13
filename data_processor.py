import openpyxl
import numpy as np
#import modules
standard_row = 200
INDEX = 0
count = 0
#read data from excel
wb = openpyxl.load_workbook('data_processor.xlsx')
sh = wb['工作表1']
columns = ['A','G','M','S','Y','AE','AK','AQ','AW','BC']

def standardlize(m, j, r):
    original_index = j * (original_row-1)/standard_row
    diff = original_index - int(original_index)
    standard_m = diff * m[int(original_index+1)][r] + (1 - diff) * m[int(original_index)][r]
    return standard_m

while INDEX < 10:
    names = sh[columns[INDEX]]
#get size
    for x in names:
        if x.value is None:
            break
        count += 1
    original_row = count
#process of the matrix
#print(sh.max_row)
#print(sh.max_column)
    raw_accel = [[0 for i in range(6)] for j in range(original_row)]

#load values
    for i in range(original_row):
        for k in range(6):
            raw_accel[i][k] = sh.cell(i+1, INDEX*6+k+1).value
    

    sta_accel = [[0 for i in range(6)] for j in range(standard_row)]
#subfunction for standardlize





    for i in range(standard_row):
        for k in range(6):
            sta_accel[i][k] = round(standardlize(raw_accel, i, k), 3)
#write back

    for i in range(standard_row):
        for k in range(6):
            sh.cell(i+1, INDEX*6+k+1, sta_accel[i][k])
    count = 0
    INDEX += 1
wb.save('data_processor.xlsx')

